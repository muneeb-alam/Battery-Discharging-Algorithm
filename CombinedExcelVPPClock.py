
        
# To run the code without primary reserve and rectify infinite loop problem
# To update the charging and discharging Price of the MarketBatteryStorage

#========================Functions============================================#

#***********************Source_MarketBatteryStorage_Functions***********************#

#This function returns updated Price of MarketBatteryStorage according to current capacity.

def MarketBatteryStorageDOD(x):
     #return (1.939E-08*(x**4) - 4.065E-06*(x**3) + 0.000326*(x**2) - 0.01141*x + 0.2219)*100
     #return -3.551E-18*(x**2)+0.05*x+28   
     return MarketBatteryStorage.DischargingPrice

# This function is used when Battery acts as source and need(load) Power is more than source Power.
# Current Capacity is updated, profit calculations for MarketBatteryStorage is done here when MarketBatteryStorage provide energy to StandardConsumingDevices and DSM.
# Dicharge factor is also defined which tells how much of MarketBatteryStorage capacity should be used while acting as....
# ....primary reserve for grid.
# No profit calculation when grid acts as need as grid operator has to pay a fixed amount to MarketBatteryStorage owner....
#... independent of Power shared. 

def MarketBatteryStorage_source_need_power_greater_than_offer_power(need,offer):
    print 'Current Capacity:', offer.PercentageCurrentCapacity    
    if need.Power >= temp_need:
        print 'case A1'
        used_power=temp_offer      
        if need.Name is 'CommonGrid' and offer.CapacityForGrid-offer.PercentageCurrentCapacity<DischargeFactor:
            pass
        elif need.Name is 'CommonGrid' and offer.CapacityForGrid-offer.PercentageCurrentCapacity>=DischargeFactor:
            return
        print offer.Name,'Profit:',offer.ProfitNormal
        offer.ProfitNormal=profit_calculation(offer,need,used_power)
        need.Power=need.Power-temp_offer
        offer.Power=offer.Power-temp_offer
        
    elif need.Power < temp_need:
        print 'case A2'
        print need.Power,temp_need
        if offer.Power-need.Power>=0:
            print 'case A21'
            used_power=need.Power
            if need.Name is 'CommonGrid' and offer.CapacityForGrid-offer.PercentageCurrentCapacity<DischargeFactor:
                pass
            elif need.Name is 'CommonGrid' and offer.CapacityForGrid-offer.PercentageCurrentCapacity>=DischargeFactor:
                return
            print offer.Name,'Profit:',offer.ProfitNormal
            offer.ProfitNormal=profit_calculation(offer,need,used_power)
            offer.Power=offer.Power-need.Power
            need.Power=0
        elif offer.Power-need.Power<0 :
            print 'case A22'
            used_power=offer.Power
            print used_power
            if need.Name is 'CommonGrid'and offer.CapacityForGrid-offer.PercentageCurrentCapacity<DischargeFactor:
                pass
            elif need.Name is 'CommonGrid' and offer.CapacityForGrid-offer.PercentageCurrentCapacity>=DischargeFactor:
                return
            print offer.Name,'Profit:',offer.ProfitNormal
            offer.ProfitNormal=profit_calculation(offer,need,used_power)
            need.Power=need.Power-offer.Power
            offer.Power=0
            
          
    offer.PercentageCurrentCapacity=((offer.PercentageCurrentCapacity*0.01*offer.UsableCapacityInKWh)-(used_power*hours/division))/offer.UsableCapacityInKWh*100                       
    print 'Capacity after Discharge', offer.PercentageCurrentCapacity
    print 'Remaining Need Power',need.Power
    print 'Remaining Offer Power',offer.Power
    
    if (offer.Power is 0 or offer.PercentageCurrentCapacity<=0):
        print 'Battery peak Power for this interval has been used'
    offer.DischargingPrice=MarketBatteryStorageDOD(100-offer.PercentageCurrentCapacity) # new Price after discharge
    offer.Price=offer.DischargingPrice    
    print 'New Discharging Price',offer.DischargingPrice
    return



# This function is used when Battery acts as source and need(load) Power is less than source Power.
# Current Capacity is updated, profit calculations for MarketBatteryStorage is done here when MarketBatteryStorage provide energy to StandardConsumingDevices and DSM.
# Dicharge factor is also defined which tells how much of MarketBatteryStorage capacity should be used while acting as....
# ....primary reserve for grid.
# No profit calculation when grid acts as need as grid operator has to pay a fixed amount to MarketBatteryStorage owner....
#... independent of Power shared. 


def MarketBatteryStorage_source_need_power_less_than_offer_power(need,offer):
     print 'Current Capacity:', offer.PercentageCurrentCapacity                    
     if need.Power >= temp_need:
         print 'case B1'
         used_power=temp_need
         if need.Name is 'CommonGrid' and offer.CapacityForGrid-offer.PercentageCurrentCapacity<DischargeFactor:
             pass
         elif need.Name is 'CommonGrid' and offer.CapacityForGrid-offer.PercentageCurrentCapacity>=DischargeFactor:
            return  
         print offer.Name,'Profit', offer.ProfitNormal
         offer.ProfitNormal=profit_calculation(offer,need,used_power)
         offer.Power=offer.Power-temp_need
         need.Power=need.Power-temp_need
         
     elif need.Power < temp_need:
         print 'case B2'
         if offer.Power-need.Power>=0:
             print 'case B21'
             used_power=need.Power
             if need.Name is 'CommonGrid'and offer.CapacityForGrid-offer.PercentageCurrentCapacity<DischargeFactor:
                 pass
             elif need.Name is 'CommonGrid' and offer.CapacityForGrid-offer.PercentageCurrentCapacity>=DischargeFactor:
                return     
             print offer.Name,'Profit=',offer.ProfitNormal  
             offer.ProfitNormal=profit_calculation(offer,need,used_power)
             offer.Power=offer.Power-need.Power
             need.Power=0
         elif offer.Power-need.Power<0:
            print 'case B22'
            used_power=offer.Power
            if need.Name is 'CommonGrid'and offer.CapacityForGrid-offer.PercentageCurrentCapacity<DischargeFactor:
                pass
            elif need.Name is 'CommonGrid' and offer.CapacityForGrid-offer.PercentageCurrentCapacity>=DischargeFactor:
                return  
            
            print offer.Name,'Profit=',offer.ProfitNormal 
            offer.ProfitNormal=profit_calculation(offer,need,used_power)
            need.Power=need.Power-offer.Power
            offer.Power=0
            
  
               
     offer.PercentageCurrentCapacity=((offer.PercentageCurrentCapacity*0.01*offer.UsableCapacityInKWh)-(used_power*hours/division))/offer.UsableCapacityInKWh*100                       
     print 'Capacity after Discharge', offer.PercentageCurrentCapacity        
     print 'Remaining Need Power',need.Power
     print 'Remaining Offer Power',offer.Power,'\n'
     if (offer.Power is 0 or offer.PercentageCurrentCapacity<=0):
        print 'Battery peak Power has been used'
     offer.DischargingPrice=MarketBatteryStorageDOD(100-offer.PercentageCurrentCapacity) # new Price after discharge
     offer.Price=offer.DischargingPrice     
     print 'New Discharging Price',offer.DischargingPrice
     return

#***********************Load_MarketBatteryStorage_Functions_End*********************#

# This function is used when Battery acts as need and need Power is less than source Power.
# Current Capacity is updated, profit calculations for MarketSolarGeneratingUnit and MarketCogenerationUnit is done here.


def MarketBatteryStorage_load_need_power_less_than_offer_power(need,offer):
        need.Price=need.ChargingPrice
        if  need.Power >= temp_need:    
            print 'case C1'
            used_power=temp_need
            if offer.Name is 'MarketSolarGeneratingUnit' or offer.Name is 'MarketCogenerationUnit':
                print 'case C11'
                offer.ProfitNormal=profit_calculation(offer,need,used_power)
                need.ProfitNormal=profit_calculation_MarketBatteryStorage_need(offer,need,used_power)
                print offer.Name, 'Profit=',offer.ProfitNormal
                print need.Name,'Profit=',need.ProfitNormal
            offer.Power=offer.Power-temp_need
            need.Power=need.Power-temp_need
            need.PercentageCurrentCapacity=need.PercentageCurrentCapacity+((temp_need*hours/division))/need.UsableCapacityInKWh*100
            print 'need CC=', need.PercentageCurrentCapacity
            
            if need.PercentageCurrentCapacity>=99 and need.PercentageCurrentCapacity<=100:
                print 'MarketBatteryStorage is fully charged now'        
            print 'Current Capacity',need.PercentageCurrentCapacity
            return need.PercentageCurrentCapacity
            
        elif need.Power < temp_need:
            print 'case C2'
            if offer.Power-need.Power>=0:
                print 'case C21'
                used_power=need.Power
                if offer.Name is 'MarketSolarGeneratingUnit' or offer.Name is 'MarketCogenerationUnit':
                    print 'case C211'                    
                    offer.ProfitNormal=profit_calculation(offer,need,used_power)
                    need.ProfitNormal=profit_calculation_MarketBatteryStorage_need(offer,need,used_power)
                    print offer.Name, 'Profit=',offer.ProfitNormal
                    print need.Name,'Profit=',need.ProfitNormal                    
                offer.Power=offer.Power-need.Power
                need.Power=0
            elif offer.Power-need.Power<0:
                print 'case C22'
                used_power=offer.Power
                if offer.Name is 'MarketSolarGeneratingUnit' or offer.Name is 'MarketCogenerationUnit' :
                    offer.ProfitNormal=profit_calculation(offer,need,used_power)
                    need.ProfitNormal=profit_calculation_MarketBatteryStorage_need(offer,need,used_power)      
                    print offer.Name, 'Profit=',offer.ProfitNormal
                    print need.Name,'Profit=',need.ProfitNormal
                need.Power=need.Power-offer.Power
                offer.Power=0
                
            need.PercentageCurrentCapacity=need.PercentageCurrentCapacity+((used_power*hours/division))/need.UsableCapacityInKWh*100
            print 'Current Capacity=',need.PercentageCurrentCapacity
            if need.PercentageCurrentCapacity>=99 and need.PercentageCurrentCapacity<=100:
                print 'MarketBatteryStorage is fully charged now'          
            print 'Current Capacity',need.PercentageCurrentCapacity
        return
        
# This function is used when Battery acts as need and need Power is less than source Power.
# Current Capacity is updated, profit calculations for MarketSolarGeneratingUnit and MarketCogenerationUnit is done here.     
        
def MarketBatteryStorage_load_need_power_more_than_offer_power(need,offer):
    need.Price=need.ChargingPrice        
    if  need.Power >= temp_need: 
        print 'case D1'
        used_power=temp_offer
        if offer.Name is 'MarketSolarGeneratingUnit' or offer.Name is 'MarketCogenerationUnit':
            print 'case D11'
            offer.ProfitNormal=profit_calculation(offer,need,used_power)
            print offer.Name, 'Profit=',offer.ProfitNormal
            need.ProfitNormal=profit_calculation_MarketBatteryStorage_need(offer,need,used_power)
            print need.Name,'Profit=',need.ProfitNormal
        need.Power=need.Power-temp_offer
        offer.Power=offer.Power-temp_offer
        
        need.PercentageCurrentCapacity=need.PercentageCurrentCapacity+((temp_offer*hours/division))/need.UsableCapacityInKWh*100
        print 'Current Capacity=',need.PercentageCurrentCapacity    
        
        if need.PercentageCurrentCapacity>=99 and need.PercentageCurrentCapacity<=100:
            print 'MarketBatteryStorage is fully charged now'          
        print 'MarketBatteryStorage current capacity',need.PercentageCurrentCapacity
        
    elif need.Power < temp_need:
        print 'case D2'
        if offer.Power-need.Power>=0:
            print 'case D21'
            used_power=need.Power
            if offer.Name is 'MarketSolarGeneratingUnit' or offer.Name is 'MarketCogenerationUnit':
                print 'case D211'                
                offer.ProfitNormal=profit_calculation(offer,need,used_power)
                print offer.Name, 'Profit=',offer.ProfitNormal
                need.ProfitNormal=profit_calculation_MarketBatteryStorage_need(offer,need,used_power)
                print need.Name,'Profit=',need.ProfitNormal  
            offer.Power=offer.Power-need.Power
            need.Power=0
            
        elif offer.Power-need.Power<0:
            print 'case D22'
            used_power=offer.Power
            if offer.Name is 'MarketSolarGeneratingUnit' or offer.Name is 'MarketCogenerationUnit':                
                print 'case D222'
                offer.ProfitNormal=profit_calculation(offer,need,used_power)
                print offer.Name, 'Profit=',offer.ProfitNormal
                need.ProfitNormal=profit_calculation_MarketBatteryStorage_need(offer,need,used_power) 
                print need.Name,'Profit=',need.ProfitNormal
            need.Power=need.Power-offer.Power
            offer.Power=0
        
        need.PercentageCurrentCapacity=need.PercentageCurrentCapacity+((used_power*hours/division))/need.UsableCapacityInKWh*100       
        if need.PercentageCurrentCapacity>=100:
            print 'MarketBatteryStorage is fully charged now'                     
        print 'Current Capacity',need.PercentageCurrentCapacity       
        print need.PercentageCurrentCapacity
    return

#**************************MarketSolarGeneratingUnit,MarketCogenerationUnit or Grid Functions********************************#

# This function is used when MarketSolarGeneratingUnit, MarketCogenerationUnit or Grid are used as source of electricity and they have less Power than needed.
# Profit calculations for theses sources is also done here.


def MarketSolarGeneratingUnit_or_KWK_or_Grid_need_power_greater_than_offer_power(need,offer):
    if need.Name is 'MarketBatteryStorage' and need.PercentageCurrentCapacity<100.0 and reserved_for_grid is 0:
        print 'case E1'
        MarketBatteryStorage_load_need_power_more_than_offer_power(need,offer)
        
    elif need.Name is not 'MarketBatteryStorage':
        print 'case E2'
        if need.Power >= temp_need:
            print 'case E21'
            if offer.Name is not 'MarketBatteryStorage':
                print 'case E211'
                used_power=temp_offer
                offer.ProfitNormal=profit_calculation(offer,need,used_power)
                print offer.Name,'Profit=',offer.ProfitNormal
            need.Power=need.Power-temp_offer
            offer.Power=offer.Power-temp_offer
        elif need.Power < temp_need:
            print 'case E22'
            if offer.Power-need.Power>=0:
                print 'case E221'
                if offer.Name is not 'MarketBatteryStorage':
                    used_power=need.Power
                    offer.ProfitNormal=profit_calculation(offer,need,used_power)
                    print 'Profit=',offer.ProfitNormal       
                offer.Power=offer.Power-need.Power
                need.Power=0
            elif offer.Power-need.Power<0:
                print 'case E222'
                if offer.Name is not 'MarketBatteryStorage':                
                    used_power=offer.Power
                    offer.ProfitNormal=profit_calculation(offer,need,used_power)
                    print 'Profit=',offer.ProfitNormal       
                need.Power=need.Power-offer.Power
                offer.Power=0          
    print 'Remaining Need Power',need.Power
    print 'Remaining Offer Power',offer.Power,'\n'
    return
    

# This function is used when MarketSolarGeneratingUnit, MarketCogenerationUnit or Grid are used as source of electricity and they have more Power than needed.
# Profit calculations for theses sources is also done here.    
        
def MarketSolarGeneratingUnit_or_KWK_or_Grid_need_power_less_than_offer_power(need,offer):      
    if need.Name is 'MarketBatteryStorage' and need.PercentageCurrentCapacity<100.0 and reserved_for_grid is 0:
        print 'case F1'
        MarketBatteryStorage_load_need_power_less_than_offer_power(need,offer)
      
    elif need.Name is not 'MarketBatteryStorage':
        print 'case F2'
        if need.Power >= temp_need:
            print 'case F21'
            if offer.Name is not 'MarketBatteryStorage':
                print 'case F211'
                used_power=temp_need
                print used_power,offer.Name,offer.Power,offer.Price,need.Name,need.Power,need.Price
                offer.ProfitNormal=profit_calculation(offer,need,used_power)
                print offer.Name,'Profit=',offer.ProfitNormal
            offer.Power=offer.Power-temp_need
            need.Power=need.Power-temp_need
        elif need.Power < temp_need:
            print 'case F22'
            if offer.Power-need.Power>=0:
                print 'case F221'
                if offer.Name is not 'MarketBatteryStorage':
                    print 'case F2211'
                    used_power=need.Power
                    offer.ProfitNormal=profit_calculation(offer,need,used_power)
                    print offer.Name,'Profit=',offer.ProfitNormal     
                offer.Power=offer.Power-need.Power
                need.Power=0
            elif offer.Power-need.Power<0:
                print 'case F222'
                if offer.Name is not 'MarketBatteryStorage':
                    used_power=offer.Power
                    offer.ProfitNormal=profit_calculation(offer,need,used_power)
                    print offer.Name,'Profit=',offer.ProfitNormal                     
                need.Power=need.Power-offer.Power
                offer.Power=0
            
    print 'Remaining Need Power',need.Power,
    print 'Remaining Offer Power',offer.Power,'\n'
    return

#***********************Profit Calculations***********************************#

# This function is used to calculate profits of MarketSolarGeneratingUnit, MarketBatteryStorage, MarketCogenerationUnit and grid.
# For the grid case, we calculate profit is calculated on fixed Price and used energy...
#...but for MarketSolarGeneratingUnit, MarketCogenerationUnit and MarketBatteryStorage, profit is calculated based on difference of need Price and source Price.

def profit_calculation(offer,need,used_power):
    if need.Name is not 'CommonGrid' and offer.Name is not 'CommonGrid':
        offer.ProfitNormal=offer.ProfitNormal+((used_power*(hours/division)*need.Price)-(used_power*(hours/division)*offer.Price))
        print offer.ProfitNormal,'PP Normal'
    elif offer.Name is 'CommonGrid':
        VPP.ProfitNormal=VPP.ProfitNormal+(used_power*(hours/division)*(-VPP.BuyingPrice-VPP.Tax+VPP.SellingPrice) )  
        print 'VPP profits',VPP.ProfitNormal
        return VPP.ProfitNormal
    return offer.ProfitNormal
    
def profit_calculation_MarketBatteryStorage_need(offer,need,used_power):
    print 'used_power',used_power
    need.ProfitNormal=need.ProfitNormal-(used_power*(hours)/division*need.Price-used_power*(hours)/division*offer.Price)
    return need.ProfitNormal     
        
# This function is used to move to next source to meet the required Power of one need
#....based on the minimum Price source. If one source is depleted, it moves on to next
#...source until either all the sources are depleted or need demand is fulfilled.
  
def Removing_Zero_Power_Sources(offers,count):
     #print 'length of offers',len(offers)
     if offers[0].Power is 0:
         if offers[1].Power is 0:
             if offers[2].Power is 0:
                 if len(offers)>3:
                     if offers[3].Power is 0:
                         count=0
                         return count
                     else:
                         count =3
                         return count
             elif offers[2].Power is not 0:
                 count=2
                 return count
         elif offers[1].Power is not 0:
            count=1
            return count
     elif offers[0].Power is not 0:
        count=0
        return count
        
# This function is used to determine the status of MarketBatteryStorage. If it is not reserved for grid
#..in next time interval, it will act as need if combined Power of MarketSolarGeneratingUnit and MarketCogenerationUnit is more than
#...combined Power of StandardConsumingDevices and DSM and in this case MarketBatteryStorage will be charged. If it is not the case, 
#...then MarketBatteryStorage will act as source and meet the need demands. If MarketBatteryStorage is reserved for grid,then
#...it will act as source or load depending on the excess or deficit of Power in grid in next time interval.
        
def status_of_MarketBatteryStorage(offers,needs):
    index=0
    for offer in offers:
        if offer.Name is 'MarketBatteryStorage' and reserved_for_grid is not 1:
            if ((MarketSolarGeneratingUnit.Power + MarketCogenerationUnit.Power)>(StandardConsumingDevices.Power + DSM.Power)):                
                offer.Price=offer.ChargingPrice
                del(offers[index])
                print 'Need Charging Price:',offer.ChargingPrice
            else:
                index_MarketBatteryStorage=index_of_MarketBatteryStorage(needs)
                offer.Price=MarketBatteryStorageDOD(100-offer.PercentageCurrentCapacity)              
                del(needs[index_MarketBatteryStorage])
                print 'Offer Discharging Price',offer.Price
            
        elif offer.Name is 'MarketBatteryStorage' and reserved_for_grid is 1:
                if CommonGrid1.Power>0.0: # positive power
                    del(offers[index])
                    print 'MarketBatteryStorage acting as a need to support the grid'
                elif CommonGrid1.Power<0.0:
                    print 'MarketBatteryStorage acting as a source to support the grid'
                    index_MarketBatteryStorage=index_of_MarketBatteryStorage(needs)
                    del(needs[index_MarketBatteryStorage])                                                                          
        else:
            index=index+1
            continue
    return

# This function gives back index of MarketBatteryStorage which is used in above function.
def index_of_MarketBatteryStorage(needs):
    index=0
    for need in needs:
        if need.Name is 'MarketBatteryStorage':
            break
        else:
            index=index+1
    return index
    
# This function defines the status of the grid based on the deficit or excess of Power.
# If grid Power is positive, it means that grid has excess Power and if it is negative, 
#...it means that grid will act as need.
    
def status_of_Grid(offers,needs):
    index=0
    for offer in offers:
        if offer.Name is 'CommonGrid':
            if offer.Power<0.0:  
                index_Grid=index_of_Grid(offers)
                #offer.Price=offer.feed_in_price
                del(offers[index_Grid])
                needs=sort_price_descending(needs) 
                offers=sort_price_ascending(offers)
                print 'Grid acting as Need with Price:',offer.Price
            elif offer.Power>=0.0:
                index_Grid=index_of_Grid(needs)
                print index_Grid
                #offer.Price=offer.draw_out_price
                del(needs[index_Grid])
                needs=sort_price_descending(needs) 
                offers=sort_price_ascending(offers)
                print 'Grid will act as a source with Price',offer.Price
        else:
            index=index+1
            continue
    return 

# This function returns the index of grid which is used in the above function.    
def index_of_Grid(needs):
    index=0
    for need in needs:
        if need.Name is 'CommonGrid':
            break
        else:
            index=index+1
    return index

# This function is used to sort needs and offers on the basis of Price in 
#...descending and ascending order respectively.

def sorting_needs_and_offers(needs,offers): 
    needs=sorted(needs, key=lambda k: k.Price, reverse = True) 
    offers=sorted(offers, key=lambda k: k.Price)
    print_needs(needs)
    print_offers(offers)
    
def sort_price_ascending(string):
    return sorted(string, key=lambda k: k.Price)
    
def sort_price_descending(string):
    return sorted(string, key=lambda k: k.Price, reverse = True)


def print_needs(needs):
    print 'Needs:'
    for need in needs:
        if need.Name is 'MarketBatteryStorage':
            print 'Name', need.Name,'Power',need.Power, 'Price', need.Price,'(current capacity in %)',need.PercentageCurrentCapacity,'usbale_capacity',need.UsableCapacityInKWh,'kWh ','Profit',need.ProfitNormal
        else:
            print 'Name',need.Name,'Power', need.Power,'Price',need.Price
    print '\n'
    
    
def print_offers(offers):
    print 'Offers:' 
    for offer in offers:
        if offer.Name is 'MarketBatteryStorage':
            print 'Name', offer.Name,'Power',offer.Power, 'Price', offer.Price,'Usable capacity',offer.UsableCapacityInKWh,'(current capacity in %)',offer.PercentageCurrentCapacity,'Profit',offer.ProfitNormal
        else:
            print 'Name', offer.Name,'Power',offer.Power, 'Price', offer.Price
            if offer.Name is not'CommonGrid':
                print 'Profit',offer.ProfitNormal
    print '\n'

def printPowers(needs,offers):
        print '\nStandardConsumingDevices Power',StandardConsumingDevicesNewPower
        print 'DSM Power',DSMNewPower
        print '\nMarketSolarGeneratingUnit Power',MarketSolarGeneratingUnitNewPower
        print 'MarketCogenerationUnit Power',MarketCogenerationNewPower        
        print '\n MarketBatteryNewCapacity',MarketBatteryNewCapacity
        print 'CommonGrid Power',CommonGridNewPower
        print '\nMarketSolarGeneratingUnit Profits',MarketSolarGeneratingUnitProfits
        print 'MarketCogenerationUnit Profits',MarketCogenerationUnitProfits
        print 'Battery Profits',MarketBatteryStorageProfits
        print  'Common Grid Profits',commonGridProfits
        
def MarketSolarGeneratingUnit_Power_Profile():
    temp=list()
    with open('MarketSolarGeneratingUnit_data.csv') as csvfile:
        readCSV=csv.reader(csvfile,delimiter=',')
        for row in readCSV:
            p=row[1]
        
            temp.append(p)
    return temp
#========================Functions_End=============================================#

#========================Needs=====================================================#

import matplotlib.pyplot as plt
import csv
from Classes import * 
from Initialization import *

# initialization of lists for profit calculations# 
MarketSolarGeneratingUnitProfits=list() 
MarketCogenerationUnitProfits=list()
MarketBatteryStorageProfits=list()
commonGridProfits=list()

#initialization of lists for Power calculations#
MarketBatteryNewCapacity=list()
MarketSolarGeneratingUnitNewPower=list()
MarketCogenerationNewPower=list()
StandardConsumingDevicesNewPower=list()
DSMNewPower=list()
CommonGridNewPower=list()

#Permanent Needs for four time intervals
#StandardConsumingDevices_Power=get_StandardConsumingDevices_Power()
#DSM_Power=get_DSM_Power()

#Permanent Offers for four time intervals
#MarketSolarGeneratingUnit_Power=get_MarketSolarGeneratingUnit_Power()
#MarketCogenerationUnitPower=get_KWK_Power()

#Conditional Needs or Offers
#Common_Grid_Power=get_Common_Grid_Power() # negative sign in Power means that grid will acts as load and postive sign means grid has excess of electricity and will act as source
#Battery_Power=get_Battery_Power() # Battery Power remains same all the time (assumption)
#Battery_Capacity=get_Battery_SOC() # Only first entry of this array will be used in code. The rest will be updated after each time interval.
#Reserve_Status=[0,0,0,0,1,1,1,1] # Reserve status for grid

#Permanent Needs for four time intervals
#StandardConsumingDevicesPower=[15,0,0,0,0,0,0,0]
#DSMPower=[5,0,0,0,0,0,0,0]

#Permanent Offers for four time intervals
#MarketSolarGeneratingUnit_Power=[0,0,0,0,0,0,0,0]
#MarketCogenerationUnitPower=[0,0,0,0,0,0,0,0]

#Conditional Needs or Offers
#Common_Grid_Power=[250,250,-500,-250,500,-1,-190,200] # negative sign in Power means that grid will acts as load and postive sign means grid has excess of electricity and will act as source
#Battery_Power=[250,250,250,250,250,250,250,250] # Battery Power remains same all the time (assumption)
#Battery_Capacity=[0,80,20,0,0,80,20,0] # Only first entry of this array will be used in code. The rest will be updated after each time interval.
#Reserve_Status=[0,0,0,0,0,0,0,0] # Reserve status for grid



from openpyxl import load_workbook
from openpyxl import Workbook




DischargeFactor=99 # How much the MarketBatteryStorage should be discharged while supporting the grid
N=10.0 # no of iterations  (to be written in floating form) 
x=4# break point
hours=1.0
division=1.0
reserved_for_grid=0
EEx.PrimaryReserveStatus=0

import os
import time


sec=0
min=0
hour=0
t=None




from openpyxl import load_workbook
from openpyxl import Workbook
#from VPPhouseholdsExcelInterfaceActualProfits import *
from Classes import *
from Initialization import *

while sec<=60:
    os.system('cls')
    #print (hour,'Hours',min, 'Minutes', sec, 'Seconds')
    print hour ,':', min, ':', sec
    if hour==2:
        break
    time.sleep(1)
    sec+=1
    if sec==60:
        min+=1
        sec=00
    if min ==60:
        hour+=1
        min=00
    
    if sec % 5: continue
    else:
        if t is None:
            t=0
        else:
            pass

        wb = load_workbook(filename = 'DatawithProfitsClock.xlsx')
        ws=wb.active
        MaxRows=ws.max_row-3
        MaxCols=ws.max_column-1
        sheet=wb.get_sheet_by_name('Data')
        a=sheet.cell(row=4,column=2)
        #print MaxRows,MaxCols
        i=t+4
        if i<MaxRows+4:
            col=2
            H1.MarketSolarGeneratingUnit.Power=sheet.cell(row=i,column=col).value
            H1.MarketCogenerationUnit.Power=sheet.cell(row=i,column=col+4).value
            CommonGrid1.Power=sheet.cell(row=i,column=col+8).value
            H1.MarketBatteryStorage.Power=sheet.cell(row=i,column=col+12).value
            H1.MarketBatteryStorage.PercentageCurrentCapacity=sheet.cell(row=i,column=col+15).value
            H1.StandardConsumingDevices.Power=sheet.cell(row=i,column=col+19).value
            H1.DSM.Power=sheet.cell(row=i,column=col+22).value
            EEx.PrimaryReserveStatus=sheet.cell(row=i,column=col+27).value
            print 'H1.MarketSolarGeneratingUnit.Power',H1.MarketSolarGeneratingUnit.Power          
            print 'H1.MarketCogenerationUnit.Power',H1.MarketCogenerationUnit.Power
            print 'CommonGrid1.Power',CommonGrid1.Power
            print 'H1.MarketBatteryStorage.Power',H1.MarketBatteryStorage.Power
            print 'H1.MarketBatteryStorage.PercentageCurrentCapacity',H1.MarketBatteryStorage.PercentageCurrentCapacity
            print 'H1.StandardConsumingDevices.Power',H1.StandardConsumingDevices.Power
            print 'H1.DSM.Power',H1.DSM.Power  
            print 'EEX PR Status',EEx.PrimaryReserveStatus
            MarketBatteryStorage=H1.MarketBatteryStorage
            DSM=H1.DSM
            MarketCogenerationUnit=H1.MarketCogenerationUnit
            MarketSolarGeneratingUnit=H1.MarketSolarGeneratingUnit
            StandardConsumingDevices=H1.StandardConsumingDevices
            t=t+1
            
        else:
            print 'No input is given'
            break
        



    try:
        print 'iteration number:',t,'\nTime',t*15,'minutes'
    except:
        pass
    #MarketSolarGeneratingUnit.Power=int(MarketSolarGeneratingUnit_Power[t])
    #MarketCogenerationUnit.Power=int(MarketCogenerationUnitPower[t])
    #StandardConsumingDevices.Power=int(StandardConsumingDevicesPower[t])
    #DSM.Power=int(DSMPower[t])
    #MarketBatteryStorage.Power=int(Battery_Power[t])
    #EEx.PrimaryReserveStatus=int(Reserve_Status[t])
    #CommonGrid1.Power=int(Common_Grid_Power[t])
    #MarketBatteryStorage.PercentageCurrentCapacity=int(Battery_Capacity[t])
    #if t is not 0: # To update the value of MarketBatteryStorage capacity after each interval 
        #MarketBatteryStorage.PercentageCurrentCapacity=int(MarketBatteryNewCapacity[t-1])  
        #print 'MarketBatteryStorage capacity',MarketBatteryStorage.PercentageCurrentCapacity
    #MarketBatteryStorage.CapacityForGrid=MarketBatteryStorage.PercentageCurrentCapacity  # reference to determine how much MarketBatteryStorage should be discharged/charged while supporting the grid
    needs=([StandardConsumingDevices,DSM,MarketBatteryStorage,CommonGrid1]) 
    offers = ([MarketSolarGeneratingUnit,MarketCogenerationUnit,MarketBatteryStorage,CommonGrid1]) 
    sorting_needs_and_offers(needs,offers)
    print "Status of Battery:",status_of_MarketBatteryStorage(offers,needs)
    print 'Status of Grid',status_of_Grid(offers,needs)
    sorting_needs_and_offers(needs,offers) # function which sorts needs on basis of Price in descending order and sorts offers on basis of Price in ascending order
    temp_offer=None
    count=0 
    print 'Primary Reserve Status', EEx.PrimaryReserveStatus,'Reserved for Grid', reserved_for_grid
    
    try:
        if MarketBatteryStorage in offers:   
            print 'BAT in OFFERS'                 
            if (-(DSM.Power*hours/division)-(StandardConsumingDevices.Power*hours/division)+(MarketSolarGeneratingUnit.Power*hours/division)+(MarketCogenerationUnit.Power*hours/division)+((MarketBatteryStorage.PercentageCurrentCapacity)/100.0*MarketBatteryStorage.UsableCapacityInKWh*hours/division))>0.0:
                print 'A'
                if ((MarketSolarGeneratingUnit.Power*hours/division)+(MarketCogenerationUnit.Power*hours/division)>(100.0-MarketBatteryStorage.PercentageCurrentCapacity)/100.0*MarketBatteryStorage.UsableCapacityInKWh*hours/division) and MarketBatteryStorage.PercentageCurrentCapacity is not 100:
                    print DSM.Power*hours/division,StandardConsumingDevices.Power*hours/division,MarketSolarGeneratingUnit.Power*hours/division,MarketCogenerationUnit.Power*hours/division,((100.0-MarketBatteryStorage.PercentageCurrentCapacity)/100.0*MarketBatteryStorage.UsableCapacityInKWh*hours/division)
                    H1.Profits=12.0*(hours/division)*((-DSM.Power*hours/division)-(StandardConsumingDevices.Power*hours/division)+(MarketSolarGeneratingUnit.Power*hours/division)+(MarketCogenerationUnit.Power*hours/division)-((100.0-MarketBatteryStorage.PercentageCurrentCapacity)/100.0*MarketBatteryStorage.UsableCapacityInKWh*hours/division))
                    print 'Battery will be fully charged'        
                    print H1.Profits, 'H1.Profits A'
                else:
                    print 'No extra feed in'
                    pass
            
            
            elif (-(DSM.Power*hours/division)-(StandardConsumingDevices.Power*hours/division)+(MarketSolarGeneratingUnit.Power*hours/division)+(MarketCogenerationUnit.Power*hours/division)+((MarketBatteryStorage.PercentageCurrentCapacity)/100.0*MarketBatteryStorage.UsableCapacityInKWh*hours/division))<0.0:
                if DSM.Price >=CommonGrid1.Price: 
                    print 'B'
                    H1.Profits=30.0*(hours/division)*((-DSM.Power*hours/division)-(StandardConsumingDevices.Power*hours/division)+(MarketSolarGeneratingUnit.Power*hours/division)+(MarketCogenerationUnit.Power*hours/division))
                elif DSM.Price<CommonGrid1.Price:
                    print 'C'                
                    H1.Profits=30.0*(hours/division)*(-(StandardConsumingDevices.Power*hours/division)+(MarketSolarGeneratingUnit.Power*hours/division)+(MarketCogenerationUnit.Power*hours/division))
                print DSM.Power*hours/division,StandardConsumingDevices.Power*hours/division,MarketSolarGeneratingUnit.Power*hours/division,MarketCogenerationUnit.Power*hours/division,((100.0-MarketBatteryStorage.PercentageCurrentCapacity)/100.0*MarketBatteryStorage.UsableCapacityInKWh*hours/division)
                print 'No extra feed in to grid'
                print H1.Profits, 'H1.Profits'
                
        elif MarketBatteryStorage in needs:
            print 'BAT in NEEDS'
            if (-(DSM.Power*hours/division)-(StandardConsumingDevices.Power*hours/division)+(MarketSolarGeneratingUnit.Power*hours/division)+(MarketCogenerationUnit.Power*hours/division)-((100-MarketBatteryStorage.PercentageCurrentCapacity)/100.0*MarketBatteryStorage.UsableCapacityInKWh*hours/division))<0.0:   
                print 'B'
                if MarketBatteryStorage.PercentageCurrentCapacity is 0:
                    if DSM.Price >=CommonGrid1.Price: 
                        print 'A'
                        H1.Profits=-30.0*(hours/division)*((-DSM.Power*hours/division)-(StandardConsumingDevices.Power*hours/division)+(MarketSolarGeneratingUnit.Power*hours/division)+(MarketCogenerationUnit.Power*hours/division))
                    elif DSM.Price<CommonGrid1.Price:
                        print 'B'                
                        H1.Profits=-30.0*(hours/division)*(-(StandardConsumingDevices.Power*hours/division)+(MarketSolarGeneratingUnit.Power*hours/division)+(MarketCogenerationUnit.Power*hours/division))
                    print DSM.Power*hours/division,StandardConsumingDevices.Power*hours/division,MarketSolarGeneratingUnit.Power*hours/division,MarketCogenerationUnit.Power*hours/division,((100.0-MarketBatteryStorage.PercentageCurrentCapacity)/100.0*MarketBatteryStorage.UsableCapacityInKWh*hours/division)
           
            elif (-(DSM.Power*hours/division)-(StandardConsumingDevices.Power*hours/division)+(MarketSolarGeneratingUnit.Power*hours/division)+(MarketCogenerationUnit.Power*hours/division)-((100-MarketBatteryStorage.PercentageCurrentCapacity)/100.0*MarketBatteryStorage.UsableCapacityInKWh*hours/division))>0.0:  
                print DSM.Power*hours/division,StandardConsumingDevices.Power*hours/division,MarketSolarGeneratingUnit.Power*hours/division,MarketCogenerationUnit.Power*hours/division,((100.0-MarketBatteryStorage.PercentageCurrentCapacity)/100.0*MarketBatteryStorage.UsableCapacityInKWh*hours/division)
                H1.Profits=12.0*(hours/division)*((-DSM.Power*hours/division)-(StandardConsumingDevices.Power*hours/division)+(MarketSolarGeneratingUnit.Power*hours/division)+(MarketCogenerationUnit.Power*hours/division)-((100.0-MarketBatteryStorage.PercentageCurrentCapacity)/100*MarketBatteryStorage.UsableCapacityInKWh*hours/division))
                print 'Battery will BE fully charged'        
                print H1.Profits, 'H1.Profits A'
                
            
                
    except:
        print 'EXCEPT'
        pass
                
    for need in needs:
        if need.Name is 'CommonGrid':
            need.Power=need.Power*-1
            print 'Common Grid Power',need.Power
        temp_need=need.Power/N
        temp_price=need.Price/N
        print 'Need Name', need.Name, 'required Power',need.Power 
        if need.Power<=0 :continue
        n=0
        iteration_var=0
        print 'iteration_var',iteration_var # to allow all the Power sources to be used to meet load demands
        
        while n < (int(N)) and need.Power>0.1 and iteration_var<(int(N+20)):
            
            print 'n=',n, 'reserved_for_grid',reserved_for_grid

            if (MarketBatteryStorage.Power >=0 and MarketBatteryStorage.PercentageCurrentCapacity>=0.0) or (MarketSolarGeneratingUnit.Power or MarketCogenerationUnit.Power or CommonGrid1.Power):
                pass
            else:
                print 'all sources are depleted'
                break
            
            if count<len(offers):
                offer=offers[count]
                
 #=====================================For Primary Reserve Settings===============================================#               

                      
            if need.Name is 'MarketBatteryStorage' and EEx.PrimaryReserveStatus is 1 and offer.Name is not 'CommonGrid':
                if need.PercentageCurrentCapacity>79:
                    print 'offer Name',offer.Name
                    print 'Primary Reserve Mode, Battery will act as reserve for grid in next interval'
                    reserved_for_grid=1
                    if offer.Name is not 'CommonGrid':
                        break
                    
                
            elif offer.Name is 'MarketBatteryStorage' and EEx.PrimaryReserveStatus is 1:
                print offer.PercentageCurrentCapacity
                if offer.PercentageCurrentCapacity<21:
                    print 'Primary Reserve Mode, Battery will act as offer for grid in next interval'
                    if reserved_for_grid is 0:
                        offer.Power=0
                        if count+1<len(offers):
                            count=count+1
                            offer=offers[count]
                        reserved_for_grid =1
                        print 'updation'
                        n=0
                        continue
            elif MarketBatteryStorage.PercentageCurrentCapacity>21 and MarketBatteryStorage.PercentageCurrentCapacity<79 and EEx.PrimaryReserveStatus is 1:
                    if (StandardConsumingDevices.Power is 0 or DSM.Power is 0) and offer.Name is 'MarketBatteryStorage' and offer.Power>0:
                        print 'reserved for grid'
                        reserved_for_grid=1
                        EEx.PrimaryReserveStatus=0
                        
 #=====================================For Primary Reserve Settings===============================================#

                        
#=================Grid reserve case when MarketBatteryStorage supplies electricity to grid======================================#
                                     
            if reserved_for_grid is 1 and need.Name is 'CommonGrid':
                if MarketBatteryStorage in needs:
                    n=n+N
                    print 'Case when MarketBatteryStorage is not available and grid needs some source'
                    break
                else:
                    offer=offers[index_of_MarketBatteryStorage(offers)]
                    print 'Offer Name is Battery for Grid', offer.Name
                    
            if reserved_for_grid is 0 and need.Name is 'CommonGrid':
                n=n+N
                print 'Reserved for Grid is 0 and Grid acting as need'
                continue
                
            if reserved_for_grid is 1 and need.Name is 'MarketBatteryStorage':
                offer=offers[index_of_Grid(offers)]
                print 'Offer Name is for MarketBatteryStorage is', offer.Name
                
            if reserved_for_grid is 1 and need.Name is 'MarketBatteryStorage':
                offer=offers[index_of_Grid(offers)]
                print 'Offer Name is for MarketBatteryStorage is', offer.Name
         
                
#=================Grid reserve case when MarketBatteryStorage supplies electricity to grid======================================#

                
            print 'Offer',offer.Name, 'Offered Power',offer.Power,'Reserve status',reserved_for_grid
            if temp_offer is None and n is 0:
                temp_offer=offer.Power/N
            elif n is 0 and temp_offer is not None:
                temp_offer=offer.Power/N
                if (int(N)==1):
                    temp_need=need.Power/1.0
            print 'temp_need',temp_need,'temp_offer',temp_offer

#=======================For Needs other than Grid==============================================================#            
            if need.Name is not 'CommonGrid':
                if temp_need-temp_offer>=0 and offer.Price<=need.Price and offer.Power>0:
                    print 'A'         
                    if offer.Name is not 'MarketBatteryStorage':
                        print 'AA'                            
                        MarketSolarGeneratingUnit_or_KWK_or_Grid_need_power_greater_than_offer_power(need,offer)
                            
                    elif offer.Name is 'MarketBatteryStorage' and offer.PercentageCurrentCapacity>0.0 and temp_need>=temp_offer and reserved_for_grid is 0:
                        print 'AB'                            
                        MarketBatteryStorage_source_need_power_greater_than_offer_power(need,offer)
               
                elif temp_need -temp_offer<0 and offer.Price <= need.Price and offer.Power>0:
                    print 'B',reserved_for_grid                   
                    if offer.Name is not 'MarketBatteryStorage':                    
                        print 'BA'                            
                        MarketSolarGeneratingUnit_or_KWK_or_Grid_need_power_less_than_offer_power(need,offer)
                        
                    elif  offer.Name is 'MarketBatteryStorage' and offer.PercentageCurrentCapacity >0.0 and temp_need<=temp_offer and reserved_for_grid is 0:       
                        print 'BB'                            
                        MarketBatteryStorage_source_need_power_less_than_offer_power(need,offer)  

#=======================For Needs other than Grid==============================================================#                        


#=======================Case when Grid acts as Need==============================================================#
         
            elif need.Name is 'CommonGrid' and reserved_for_grid is 1 and offer.Name is 'MarketBatteryStorage': 
                print 'C'
                if temp_need-temp_offer>=0 and offer.Power>0 :
                    print 'C'
                    if offer.PercentageCurrentCapacity>0.0 and reserved_for_grid is 1:
                        print 'CA'                            
                        MarketBatteryStorage_source_need_power_greater_than_offer_power(need,offer)
                        n=n+1
                
                elif temp_need-temp_offer<0 and offer.Power>0 :
                    print 'D'
                    if offer.PercentageCurrentCapacity>0.0 and reserved_for_grid is 1:
                        print 'DA'                            
                        MarketBatteryStorage_source_need_power_less_than_offer_power(need,offer)  
                        n=n+1
#=======================Case when Grid acts as Need==============================================================#                        
             
            
            if offer.Name is 'MarketBatteryStorage':
                if offer.Power>0 and offer.PercentageCurrentCapacity>0:
                    if need.Name is not 'CommonGrid' and offer.Price<need.Price:
                        offer.Power=offer.Power
                    elif need.Name is 'CommonGrid':
                        continue
                else:
                    if need.Name is not 'CommonGrid':
                        print 'Making MarketBatteryStorage Power zero to use another source'
                        offer.Power=0
                    elif need.Name is 'CommonGrid':
                        reserved_for_grid=0
                        need.Power=need.Power*-1
                        break             
                    
#=======================Case when Grid acts as Source for Battery during high generation==============================================================#                

            if offer.Name is 'CommonGrid' and offer.Power>=0:
                print 'E'
                if reserved_for_grid is 1 and need.Name is 'MarketBatteryStorage' and need.PercentageCurrentCapacity<100:
                    print 'Battery acting as load for Grid'
                    if temp_need-temp_offer>=0:
                        print 'EA'
                        MarketBatteryStorage_load_need_power_more_than_offer_power(need,offer)
                        n=n+1
                        continue
                        
                    elif  temp_need-temp_offer<0:
                        print 'EB'
                        MarketBatteryStorage_load_need_power_less_than_offer_power(need,offer)
                        n=n+1
                        continue
                    
#=======================Case when Grid acts as Source for Battery during high generation==============================================================#                
                          
                    
#=======================No Power sharing if Price of offer is more than need Price==============================================================#                
                                                            
            elif offer.Price>need.Price and need.Name is not 'CommonGrid':
                print 'Offered Price', offer.Price,'of', offer.Name,'is greater than need Price',need.Price, 'of', need.Name
                n=n+1
#=======================No Power sharing if Price of offer is more than need Price==============================================================#                

#=========================To move onto next need, if current need is fulfilled======================================#

            if (need.Power<=0.001):
                need.Power=0
                print need.Name, 'Power demand is satisfied now for this interval \n'
                print_needs(needs)
                print_offers(offers) 
                #if need.Name is 'Common_Grid':
                    #reserved_for_grid=0 # resetting the grid status   
                if offer.Name is 'MarketBatteryStorage':
                    print 'Check reserve status'
                    continue  
                break
#=========================To move onto next need, if current need is fulfilled======================================#
             
             
#=========================To move to next offer to satisfy one need======================================#                        
            if (offer.Power<0.001):
                offer.Power=0
                print offer.Name,'is delpleted \n'
                if count+1<len(offers):
                    count=count+1
                else: 
                    break
                offers = sort_price_ascending(offers)

                                        
            offers=sort_price_ascending(offers)
            prev_offer=offer.Name
            print 'COUNT',count
            count=Removing_Zero_Power_Sources(offers,count) 
            print 'COUNT',count
                
            if count is None:count=0
            if offers[count].Name is not prev_offer:
                iteration_var=iteration_var+1
                offer=offers[count]
                print 'new offer', offer.Name
                n=0
                continue
                
            if offers[count].Name is prev_offer:
                n=n+1
            print_needs(needs)
            print_offers(offers) 
            
#=========================To move to next offer to satisfy one need======================================#
            
    print 'End of iteration',t
    if EEx.PrimaryReserveStatus is 1:
        print 'CHANGE'
        reserved_for_grid =1
    elif EEx.PrimaryReserveStatus is 0:
        reserved_for_grid=0
        


    MarketSolarGeneratingUnitProfits.append(MarketSolarGeneratingUnit.ProfitNormal)
    MarketCogenerationUnitProfits.append(MarketCogenerationUnit.ProfitNormal)
    MarketBatteryStorageProfits.append(MarketBatteryStorage.ProfitNormal)    
    commonGridProfits.append(CommonGrid1.ProfitNormal)
    MarketBatteryNewCapacity.append(MarketBatteryStorage.PercentageCurrentCapacity)
    MarketSolarGeneratingUnitNewPower.append(MarketSolarGeneratingUnit.Power)
    MarketCogenerationNewPower.append(MarketCogenerationUnit.Power)
    StandardConsumingDevicesNewPower.append(StandardConsumingDevices.Power)
    DSMNewPower.append(DSM.Power)
    CommonGridNewPower.append(CommonGrid1.Power)        
    printPowers(needs,offers)
    
    
    try:
        print 'H1.Profits',H1.Profits
    except:
        pass
    print 'Primary Reserve Status', EEx.PrimaryReserveStatus,'Reserved for Grid', reserved_for_grid
    
    col=5
    print 'I',i, 'col',col
    sheet.cell(row=i,column=col).value=MarketSolarGeneratingUnit.ProfitNormal
    sheet.cell(row=i,column=col+4).value=MarketCogenerationUnit.ProfitNormal
    sheet.cell(row=i,column=col+8).value=CommonGrid1.ProfitNormal
    sheet.cell(row=i,column=col+15).value=MarketBatteryStorage.ProfitNormal
    try:
        sheet.cell(row=i,column=col+22).value=H1.Profits
    except:
        pass
    sheet.cell(row=i,column=col+23).value=VPP.ProfitNormal

    wb.save('DatawithProfitsClock.xlsx')   
    if t is x:break


print VPP.ProfitNormal
print H1.Profits