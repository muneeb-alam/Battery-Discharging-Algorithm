class TimeSeries:
    
    def __init__(self,Hour,Price,XFactor,AgeingPrice,Cost):
        self.Hour=Hour
        self.Price=Price
        self.XFactor=XFactor
        self.AgeingPrice=AgeingPrice
        self.Cost=Cost
 
lst=list()
Templst=list()
XFactorslst=list()
XFactorlst=list()

from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook(filename = 'MarketPrices.xlsx')
ws=wb.active
MaxRows=ws.max_row-8
MaxCols=ws.max_column-1
sheet=wb.get_sheet_by_name('Data') 

j=9
while j<=MaxRows+8:
    T=TimeSeries(sheet.cell(row=j,column=2).value,sheet.cell(row=j,column=3).value,1,0.5,0)
    lst.append(T)
    j+=1
   
def Ageing_Factor(AF):
    if AF is 0:
        return 0
    elif AF is 1:
        return 0.5
    elif AF is 2:
        return 1.4
    elif AF is 3:
        return 2.5
    elif AF is 4:
        return 4
    else:
        return 4.5
       
Demand=3 
BatteryCapacity=15.0 #in kWh
Length=len(lst)
k=0
TotalCost=list()
hours=MaxRows
while k <=Length:
    Sum=list()
    i=0
    print 'iteration number',k
    if k is 0:
        x=9
        for element in lst:
            print 'x',x,element.Price,Demand,BatteryCapacity/hours,element.AgeingPrice
            element.Cost=element.Price*(Demand-BatteryCapacity/hours)+element.AgeingPrice
            sheet.cell(row=x,column=4).value=element.Cost
            x+=1
            XFactorlst.append(element.XFactor)
            wb.save('MarketPrices.xlsx')
        lst=sorted(lst, key=lambda k: k.Cost, reverse = True)
        for element in lst:
            Sum.append(element.Cost)
            print 'Hour',element.Hour,'Price',element.Price,'XFactor',element.XFactor,'Cost',element.Cost
        for element in Templst:
            Sum.append(element.Cost)
        print 'TotalCost',sum(Sum)
        Sum=list()
        XFactorslst.append(XFactorlst)
        XFactorlst=list()
    
    print 'Before optimization'
           
    if lst[i].Cost-lst[len(lst)-1].Cost>0:
        print 'CONDITION SATISFIED'
        lst[i].XFactor+=1
        lst[i].AgeingPrice=Ageing_Factor(lst[i].XFactor)
        lst[i].Cost=lst[i].Price*(Demand-(BatteryCapacity/hours)*lst[i].XFactor)+lst[i].AgeingPrice
        lst[len(lst)-1].XFactor-=1
        lst[len(lst)-1].AgeingPrice=Ageing_Factor(lst[len(lst)-1].XFactor)
        lst[len(lst)-1].Cost=lst[len(lst)-1].Price*(Demand-(BatteryCapacity/hours)*lst[len(lst)-1].XFactor)+lst[len(lst)-1].AgeingPrice
        for element in lst:
            Sum.append(element.Cost)
        for element in Templst:
            Sum.append(element.Cost)
        print 'TotalCost after optimization',sum(Sum)
        print 'After optimization'
        for element in lst:
            print 'Hour',element.Hour,'Price',element.Price,'XFactor',element.XFactor,'Cost',element.Cost
    
        if lst[len(lst)-1].XFactor is 0:
            Templst.append(lst.pop(len(lst)-1)) 
            
        for element in lst:
            Templst.append(element)
        
        Templst=sorted(Templst, key=lambda k: k.Hour)
        for element in Templst:
            XFactorlst.append(element.XFactor)
        XFactorslst.append(XFactorlst)
        print 'XFactorslst',XFactorslst
        print 'Templst',Templst
        for element in lst:
            if element in Templst:
                #print 'del'
                Templst.pop(Templst.index(element))
                
        print 'Templst',len(Templst)
        print 'lst',len(lst)
        lst=sorted(lst, key=lambda k: k.Cost, reverse = True)
        for element in lst:
            print 'Hour',element.Hour,'Price',element.Price,'XFactor',element.XFactor,'Cost',element.Cost
        TotalCost.append(sum(Sum))
        Sum=list()
        XFactorlst=list()
            
    else:
        break
    if k is 15: break
    k+=1

print '\n Templist'
for element in lst:
    Templst.append(element)
    
Templst=sorted(Templst, key=lambda k: k.Hour)
 
for element in TotalCost:
    print element
    
               
for l in TotalCost:
    print 'Total Costs',l
    
x=0

index=TotalCost.index(min(TotalCost))
print index,min(TotalCost)

XFactorlst=XFactorslst[index+1]
print XFactorlst
i=0
for element in Templst:
    element.XFactor=XFactorlst[i]
    element.AgeingPrice=Ageing_Factor(element.XFactor)
    element.Cost=element.Price*(Demand-(BatteryCapacity/hours)*element.XFactor)+element.AgeingPrice
    i+=1

for element in Templst:
        print 'Hour',element.Hour,'Price',element.Price,'XFactor',element.XFactor,'Cost',element.Cost
        sheet.cell(row=x+9,column=5).value=element.XFactor
        sheet.cell(row=x+9,column=6).value=element.Cost
        x=x+1
               
wb.save('MarketPrices.xlsx')

print 'SSS'
for element in XFactorslst:
    print element
