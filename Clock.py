import os
import time


sec=0
min=0
hour=0
t=None




from openpyxl import load_workbook
from openpyxl import Workbook
#from VPPhouseholdsExcelInterfaceActualProfits import *
from Classes import *
from Initialization import *

while sec<=60:
    os.system('cls')
    #print (hour,'Hours',min, 'Minutes', sec, 'Seconds')
    print hour ,':', min, ':', sec
    if hour==2:
        break
    time.sleep(1)
    sec+=1
    if sec==60:
        min+=1
        sec=00
    if min ==60:
        hour+=1
        min=00
    
    if sec % 2: continue
    else:
        if t is None:
            t=0
        else:
            pass

        wb = load_workbook(filename = 'DatawithProfitsClock.xlsx')
        ws=wb.active
        MaxRows=ws.max_row-3
        MaxCols=ws.max_column-1
        sheet=wb.get_sheet_by_name('Data')
        a=sheet.cell(row=4,column=2)
        #print MaxRows,MaxCols
        i=t+4
        if i<MaxRows+4:
            col=2
            H1.MarketSolarGeneratingUnit.Power=sheet.cell(row=i,column=col).value
            H1.MarketCogenerationUnit.Power=sheet.cell(row=i,column=col+4).value
            CommonGrid1.Power=sheet.cell(row=i,column=col+8).value
            H1.MarketBatteryStorage.Power=sheet.cell(row=i,column=col+12).value
            H1.MarketBatteryStorage.PercentageCurrentCapacity=sheet.cell(row=i,column=col+15).value
            H1.StandardConsumingDevices.Power=sheet.cell(row=i,column=col+19).value
            H1.DSM.Power=sheet.cell(row=i,column=col+22).value
            EEx.PrimaryReserveStatus=sheet.cell(row=i,column=col+27).value
            print 'H1.MarketSolarGeneratingUnit.Power',H1.MarketSolarGeneratingUnit.Power          
            print 'H1.MarketCogenerationUnit.Power',H1.MarketCogenerationUnit.Power
            print 'CommonGrid1.Power',CommonGrid1.Power
            print 'H1.MarketBatteryStorage.Power',H1.MarketBatteryStorage.Power
            print 'H1.MarketBatteryStorage.PercentageCurrentCapacity',H1.MarketBatteryStorage.PercentageCurrentCapacity
            print 'H1.StandardConsumingDevices.Power',H1.StandardConsumingDevices.Power
            print 'H1.DSM.Power',H1.DSM.Power  
            print 'EEX PR Status',EEx.PrimaryReserveStatus
            MarketBatteryStorage=H1.MarketBatteryStorage
            DSM=H1.DSM
            MarketCogenerationUnit=H1.MarketCogenerationUnit
            MarketSolarGeneratingUnit=H1.MarketSolarGeneratingUnit
            StandardConsumingDevices=H1.StandardConsumingDevices
            t=t+1
            
        else:
            print 'No input is given'
            break